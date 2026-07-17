#!/usr/bin/env python3
"""
Stage C of the YouTube recipe-import pipeline (see scripts/youtube/README.md).

Joins the Claude enrichment (Stage B) back to the raw video metadata (Stage A),
tries to match each dish against the INDB workbook (the same nutrition source
server/scripts/indb/ uses) for VERIFIED per-serving nutrition, and only falls
back to the model's own estimated_* fields when no confident match exists --
recorded via `nutrition_source: verified|estimated` so the fallback is never
silently treated as ground truth. Validates required fields, dedups by slug,
and appends accepted dishes to recipes.json.

  python scripts/youtube/merge.py --dry-run     # counts only, writes nothing
  python scripts/youtube/merge.py               # append accepted dishes to recipes.json

Then reseed with the existing idempotent script:
  php scripts/seed.php
"""
import argparse
import difflib
import glob
import json
import os
import re
import sys
import urllib.request

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SEED_DIR = os.path.join(HERE, "..", "..", "database", "seed")
WORK_DIR = os.path.join(SEED_DIR, "youtube")
RAW_DIR = os.path.join(WORK_DIR, "raw")
ENRICHED_DIR = os.path.join(WORK_DIR, "enriched")
RECIPES = os.path.join(SEED_DIR, "recipes.json")

# Same INDB workbook server/scripts/indb/extract.py downloads/caches -- reused
# here as a nutrition lookup, not just the pre-filtered indb_candidates.json,
# since a YouTube dish might match a dessert or an already-catalogued dish
# that INDB's own category filtering excludes.
INDB_XLSX = os.path.join(SEED_DIR, "indb", "Anuvaad_INDB_2024.11.xlsx")
INDB_XLSX_URL = "https://www.anuvaad.org.in/wp-content/uploads/2020/07/Anuvaad_INDB_2024.11.xlsx"

VIDEO_ID_RE = re.compile(r"[?&]v=([A-Za-z0-9_-]{11})")

MEAL = {"breakfast", "brunch", "lunch", "dinner", "snack"}
FOOD = {"veg", "egg", "nonveg"}
CAT = {"main", "bread", "rice", "snack", "beverage", "dessert"}
DIFF = {"easy", "medium", "hard"}


def norm(s):
    return re.sub(r"[^a-z ]", "", re.sub(r"\(.*?\)", "", str(s).lower())).strip()


def slugify(name):
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", norm(name))).strip("-")


def channel_slug(handle):
    return re.sub(r"[^a-z0-9]+", "-", handle.lstrip("@").lower()).strip("-")


def load_indb_nutrition():
    """Full-workbook name -> verified per-serving nutrition, bucketed by first
    word for cheap fuzzy matching. Applies the same Atwater QA check as
    indb/extract.py (macros must actually explain the calorie count) but NOT
    its category-exclusion/dedup filters -- we want a match even for a
    dessert or a dish already in recipes.json, since we're borrowing nutrition
    truth, not deciding whether to add an INDB row as its own recipe.
    """
    if not os.path.exists(INDB_XLSX):
        print(f"Downloading INDB workbook -> {INDB_XLSX}")
        os.makedirs(os.path.dirname(INDB_XLSX), exist_ok=True)
        urllib.request.urlretrieve(INDB_XLSX_URL, INDB_XLSX)

    wb = openpyxl.load_workbook(INDB_XLSX, read_only=True, data_only=True)
    rows = wb[wb.sheetnames[0]].iter_rows(values_only=True)
    H = {h: i for i, h in enumerate(next(rows))}

    def g(r, k):
        v = r[H[k]]
        return v if v is not None else 0

    lookup, by_first_word = {}, {}
    for r in rows:
        name = str(g(r, "food_name")).strip()
        if not name:
            continue
        kcal, prot, carb, fat = (g(r, k) for k in
                                  ("unit_serving_energy_kcal", "unit_serving_protein_g",
                                   "unit_serving_carb_g", "unit_serving_fat_g"))
        if kcal <= 0:
            continue
        atwater = prot * 4 + carb * 4 + fat * 9
        if abs(atwater - kcal) > 0.25 * kcal:      # same QA check as indb/extract.py
            continue

        nn = norm(name)
        if nn in lookup:
            continue
        vitc, vita, fol = (g(r, k) for k in
                            ("unit_serving_vitc_mg", "unit_serving_vita_ug", "unit_serving_folate_ug"))
        lookup[nn] = {
            "calories": round(kcal), "protein_g": round(prot), "carbs_g": round(carb),
            "fat_g": round(fat), "fiber_g": round(g(r, "unit_serving_fibre_g")),
            "calcium_mg": round(g(r, "unit_serving_calcium_mg")),
            "vitamin_score": min(5, max(1, 2 + (vitc > 5) + (vita > 50) + (fol > 30))),
        }
        first = nn.split(" ")[0] if nn else ""
        by_first_word.setdefault(first, []).append(nn)

    return lookup, by_first_word


def match_nutrition(name, lookup, by_first_word, threshold):
    nn = norm(name)
    if nn in lookup:
        return lookup[nn], 1.0
    first = nn.split(" ")[0] if nn else ""
    best, best_ratio = None, 0.0
    for cand in by_first_word.get(first, []):
        ratio = difflib.SequenceMatcher(None, nn, cand).ratio()
        if ratio > best_ratio:
            best, best_ratio = cand, ratio
    if best and best_ratio >= threshold:
        return lookup[best], best_ratio
    return None, 0.0


def load_enriched(source):
    if source == "smoke":
        return json.load(open(os.path.join(WORK_DIR, "_smoke.json"), encoding="utf-8"))
    if not os.path.isdir(ENRICHED_DIR):
        sys.exit(f"No enriched chunks at {ENRICHED_DIR} -- run extract.py first.")
    merged = {}
    for f in sorted(os.listdir(ENRICHED_DIR)):
        if f.startswith("chunk_") and f.endswith(".json"):
            merged.update(json.load(open(os.path.join(ENRICHED_DIR, f), encoding="utf-8")))
    return merged


def load_raw_by_id():
    raw = {}
    for path in glob.glob(os.path.join(RAW_DIR, "*", "*.json")):
        v = json.load(open(path, encoding="utf-8"))
        raw[v["video_id"]] = v
    return raw


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--source", choices=["chunks", "smoke"], default="chunks")
    ap.add_argument("--nutrition-threshold", type=float, default=0.85,
                     help="minimum fuzzy name-match ratio to trust an INDB nutrition match")
    args = ap.parse_args()

    enriched = load_enriched(args.source)
    raw_by_id = load_raw_by_id()
    existing = json.load(open(RECIPES, encoding="utf-8"))
    existing_slugs = {r["slug"] for r in existing}
    # load_enriched() re-reads every chunk_NN.json ever produced (they're never
    # deleted), so without this check every run would re-append every
    # already-merged video under a new slug-collision suffix (e.g. "-2", "-3")
    # instead of skipping it -- that's exactly what inflated recipes.json with
    # thousands of duplicate rows before this fix.
    already_merged_video_ids = set()
    for r in existing:
        m = VIDEO_ID_RE.search(r.get("video_url") or "")
        if m:
            already_merged_video_ids.add(m.group(1))
    indb_lookup, indb_by_first_word = load_indb_nutrition()

    accepted, dropped, rejected, already_merged = [], [], [], []
    seen = set()
    verified_count = estimated_count = 0

    for vid, e in enriched.items():
        if vid in already_merged_video_ids:
            already_merged.append(vid)
            continue
        video = raw_by_id.get(vid)
        if not video:
            rejected.append((vid, "no raw video metadata"))
            continue
        if not e.get("is_recipe"):
            dropped.append((vid, e.get("reason", "not a recipe")))
            continue

        # Common extraction mix-up: a dish_category value ('dessert'/'beverage') put in
        # meal_type instead. These are otherwise-good recipes -- recover them by mapping
        # to the closest real meal slot rather than rejecting.
        if e.get("meal_type") in ("dessert", "beverage") and e.get("meal_type") not in MEAL:
            e["meal_type"] = "snack"

        if (e.get("meal_type") not in MEAL or e.get("food_type") not in FOOD
                or e.get("dish_category") not in CAT or e.get("difficulty") not in DIFF
                or not e.get("name") or not e.get("ingredients") or not e.get("instructions")):
            rejected.append((vid, "invalid/missing fields"))
            continue

        name = e["name"].strip()
        base_slug = f"{slugify(name)}-{channel_slug(video['channel_handle'])}"
        slug, n = base_slug, 2
        while slug in existing_slugs or slug in seen:
            slug = f"{base_slug}-{n}"
            n += 1
        seen.add(slug)

        nutrition, ratio = match_nutrition(name, indb_lookup, indb_by_first_word, args.nutrition_threshold)
        if nutrition:
            nutrition_source = "verified"
            verified_count += 1
        else:
            nutrition_source = "estimated"
            estimated_count += 1
            nutrition = {
                "calories": e["estimated_calories"], "protein_g": e["estimated_protein_g"],
                "carbs_g": e["estimated_carbs_g"], "fat_g": e["estimated_fat_g"],
                "fiber_g": e["estimated_fiber_g"], "calcium_mg": e["estimated_calcium_mg"],
                "vitamin_score": 2,
            }

        accepted.append((vid, {
            "slug": slug,
            "name": name,
            "cuisine": (e.get("cuisine") or "Indian").strip()[:64],
            "meal_type": e["meal_type"],
            "food_type": e["food_type"],
            "dish_category": e["dish_category"],
            "servings": int(e.get("servings") or 2),
            "calories": nutrition["calories"], "protein_g": nutrition["protein_g"],
            "carbs_g": nutrition["carbs_g"], "fat_g": nutrition["fat_g"],
            "fiber_g": nutrition["fiber_g"], "calcium_mg": nutrition["calcium_mg"],
            "vitamin_score": nutrition["vitamin_score"],
            "nutrition_source": nutrition_source,
            "contains_egg": int(e.get("contains_egg") or 0),
            "contains_onion": int(e.get("contains_onion") or 0),
            "contains_garlic": int(e.get("contains_garlic") or 0),
            "is_kid_friendly": int(e.get("is_kid_friendly") or 0),
            "is_high_protein": int(e.get("is_high_protein") or 0),
            "is_low_carb": int(e.get("is_low_carb") or 0),
            "is_weight_loss": int(e.get("is_weight_loss") or 0),
            "ingredients": [str(x).strip().lower() for x in e["ingredients"] if str(x).strip()][:20],
            "instructions": e["instructions"].strip(),
            "prep_time_min": int(e.get("prep_time_min") or 20),
            "difficulty": e["difficulty"],
            "image_url": f"https://i.ytimg.com/vi/{vid}/hqdefault.jpg",
            "video_url": video["url"],
            "source_channel": video["channel"],
        }))

    # Cross-channel duplicate resolution: the same dish (by normalized name) can come
    # from more than one channel/video in a batch like this. Keep only the one from the
    # most-viewed video; drop the rest rather than adding every channel's take on it.
    by_name = {}
    for vid, rec in accepted:
        by_name.setdefault(norm(rec["name"]), []).append((vid, rec))

    final_accepted = []
    dupes_dropped = []
    for group in by_name.values():
        if len(group) == 1:
            final_accepted.append(group[0][1])
            continue
        group.sort(key=lambda vr: raw_by_id.get(vr[0], {}).get("view_count", 0), reverse=True)
        winner_vid, winner_rec = group[0]
        final_accepted.append(winner_rec)
        for vid, rec in group[1:]:
            dupes_dropped.append((
                rec["slug"], raw_by_id.get(vid, {}).get("view_count", 0),
                winner_rec["slug"], raw_by_id.get(winner_vid, {}).get("view_count", 0),
            ))

    print(f"videos={len(enriched)} | already_merged={len(already_merged)} "
          f"accepted={len(accepted)} dropped(not-a-recipe)={len(dropped)} rejected={len(rejected)}")
    print(f"  nutrition: verified(INDB match)={verified_count} estimated(AI fallback)={estimated_count}")
    print(f"  cross-channel duplicates dropped (kept most-viewed): {len(dupes_dropped)} "
          f"| final={len(final_accepted)}")
    if dropped[:5]:
        print("  sample drops:", "; ".join(f"{v}({r})" for v, r in dropped[:5]))
    if rejected[:5]:
        print("  sample rejects:", "; ".join(f"{v}({r})" for v, r in rejected[:5]))
    if dupes_dropped[:5]:
        print("  sample duplicate drops:", "; ".join(
            f"{s}({v} views) kept {ws}({wv} views)" for s, v, ws, wv in dupes_dropped[:5]))

    if args.dry_run:
        print(f"\n[dry-run] would grow recipes.json {len(existing)} -> {len(existing) + len(final_accepted)}")
        return
    if not final_accepted:
        print("\nNothing to append.")
        return

    # Append in the file's existing one-object-per-line compact style (minimal git diff),
    # rather than re-serialising the whole array -- same approach as indb/merge.py.
    text = open(RECIPES, encoding="utf-8").read()
    cut = text.rstrip().rfind("]")
    head = text[:cut].rstrip()
    block = ",\n" + ",\n".join("  " + json.dumps(a, ensure_ascii=False) for a in final_accepted) + "\n"
    open(RECIPES, "w", encoding="utf-8").write(head + block + "]\n")

    total = len(existing) + len(final_accepted)
    print(f"\nrecipes.json now has {total} dishes (+{len(final_accepted)}). Run: php scripts/seed.php")


if __name__ == "__main__":
    sys.exit(main())
