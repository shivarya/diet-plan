#!/usr/bin/env python3
"""
Stage A of the INDB enrichment pipeline (see scripts/indb/README.md).

Deterministic, no AI. Reads the Indian Nutrient Databank workbook, keeps the
Atwater-consistent savoury dishes that aren't already in our catalogue, maps the
authoritative per-serving nutrition onto our schema, and writes a candidate list
for Stage B (enrich.py) to fill in ingredients/method/flags.

  python scripts/indb/extract.py

Output: server/database/seed/indb/indb_candidates.json
"""
import json
import os
import re
import sys
import urllib.request

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SEED_DIR = os.path.join(HERE, "..", "..", "database", "seed")
WORK_DIR = os.path.join(SEED_DIR, "indb")
XLSX = os.path.join(WORK_DIR, "Anuvaad_INDB_2024.11.xlsx")
XLSX_URL = "https://www.anuvaad.org.in/wp-content/uploads/2020/07/Anuvaad_INDB_2024.11.xlsx"
RECIPES = os.path.join(SEED_DIR, "recipes.json")
OUT = os.path.join(WORK_DIR, "indb_candidates.json")

# Drop drinks, sweets, condiments, and non-dish rows by name keyword.
EXCLUDE = [
    "tea", "coffee", "juice", "sharbat", "sherbet", "lassi", "shake", "smoothie",
    "kheer", "halwa", "jamun", "jalebi", "barfi", "burfi", "laddu", "ladoo",
    "payasam", "basundi", "rabri", "rabdi", "ice cream", "kulfi", "custard",
    "cake", "biscuit", "cookie", "candy", "toffee", "sweet", "syrup", "soda",
    "cola", "wine", "beer", " milk", "buttermilk", "water", "mocktail", "falooda",
    "peda", "mysore pak", "gulkand", "chikki", "jam ", "pickle", "chutney",
    "papad", "sugar", "mishti", "sandesh", "modak", "malpua", "squash", "thandai",
    "gajak", "petha", "kalakand", "ghee", "butter ", "honey", "jaggery", "sauce only",
    "cocoa", "nog", "drink", "beverage", "mousse", "pudding", "trifle", "smooth",
]
# First-guess food_type (Haiku corrects in Stage B). "egg" only if no meat present.
MEAT = ["chicken", "mutton", "lamb", "fish", "prawn", "shrimp", "crab", "meat",
        "beef", "pork", "keema", "kheema", "goat", "ham", "bacon", "sausage", "tikka"]
EGG = ["egg", "omelette", "omlet", "akuri", "anda"]
# First-guess meal_type.
BREAKFAST = ["dosa", "idli", "upma", "poha", "paratha", "thepla", "chilla", "cheela",
             "uttapam", "pongal", "dalia", "oats", "omelette", "appam", "adai",
             "pesarattu", "paniyaram", "vada", "dhokla", "sandwich", "muesli",
             "sabudana", "cornflakes", "porridge", "pancake", "waffle"]
DINNER = ["soup", "shorba", "tikka", "kabab", "kebab", "grilled", "roast", "salad",
          "stew", "tandoori", "broth"]


def norm(s):
    return re.sub(r"[^a-z ]", "", re.sub(r"\(.*?\)", "", str(s).lower())).strip()


def slugify(name):
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", norm(name))).strip("-") + "-indb"


def main():
    if not os.path.exists(XLSX):
        print(f"Downloading INDB workbook -> {XLSX}")
        os.makedirs(WORK_DIR, exist_ok=True)
        urllib.request.urlretrieve(XLSX_URL, XLSX)

    existing = json.load(open(RECIPES, encoding="utf-8"))
    exist_norm = {norm(r["name"]) for r in existing}
    exist_slug = {r["slug"] for r in existing}

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    rows = wb[wb.sheetnames[0]].iter_rows(values_only=True)
    H = {h: i for i, h in enumerate(next(rows))}

    def g(r, k):
        v = r[H[k]]
        return v if v is not None else 0

    candidates = []
    seen_slugs = set()
    total_savory = rejected_atwater = 0
    for r in rows:
        name = str(g(r, "food_name")).strip()
        nl = name.lower()
        if not name or any(e in nl for e in EXCLUDE):
            continue
        kcal = g(r, "unit_serving_energy_kcal")
        prot = g(r, "unit_serving_protein_g")
        carb = g(r, "unit_serving_carb_g")
        fat = g(r, "unit_serving_fat_g")
        fib = g(r, "unit_serving_fibre_g")
        cal = g(r, "unit_serving_calcium_mg")
        if not (80 <= kcal <= 600) or prot < 3:
            continue
        total_savory += 1
        atwater = prot * 4 + carb * 4 + fat * 9
        if kcal <= 0 or abs(atwater - kcal) > 0.25 * kcal:   # QA: macros must explain kcal
            rejected_atwater += 1
            continue

        nn = norm(name)
        if nn in exist_norm:
            continue
        slug = slugify(name)
        if slug in exist_slug or slug in seen_slugs:
            continue
        seen_slugs.add(slug)

        if any(x in nl for x in MEAT):
            food = "nonveg"
        elif any(x in nl for x in EGG):
            food = "egg"
        else:
            food = "veg"
        meal = ("breakfast" if any(x in nl for x in BREAKFAST)
                else "dinner" if any(x in nl for x in DINNER) else "lunch")

        vitc = g(r, "unit_serving_vitc_mg")
        vita = g(r, "unit_serving_vita_ug")
        fol = g(r, "unit_serving_folate_ug")
        vscore = min(5, max(1, 2 + (vitc > 5) + (vita > 50) + (fol > 30)))

        candidates.append({
            "slug": slug,
            "name": name,
            "meal_type_guess": meal,
            "food_type_guess": food,
            # authoritative INDB nutrition (locked — Stage C re-applies these verbatim)
            "calories": round(kcal),
            "protein_g": round(prot),
            "carbs_g": round(carb),
            "fat_g": round(fat),
            "fiber_g": round(fib),
            "calcium_mg": round(cal),
            "vitamin_score": vscore,
        })

    os.makedirs(WORK_DIR, exist_ok=True)
    json.dump(candidates, open(OUT, "w", encoding="utf-8"), indent=2, ensure_ascii=False)

    pct = 100 * rejected_atwater // max(1, total_savory)
    print(f"Savoury in range: {total_savory} | rejected by Atwater (bad data): "
          f"{rejected_atwater} ({pct}%) | candidates (deduped): {len(candidates)}")
    by_food, by_meal = {}, {}
    for c in candidates:
        by_food[c["food_type_guess"]] = by_food.get(c["food_type_guess"], 0) + 1
        by_meal[c["meal_type_guess"]] = by_meal.get(c["meal_type_guess"], 0) + 1
    print("  food_type guess:", by_food)
    print("  meal_type guess:", by_meal)
    print("  sample:", ", ".join(c["name"] for c in candidates[:8]))
    print(f"Wrote {len(candidates)} -> {os.path.relpath(OUT, os.path.join(HERE, '..', '..'))}")


if __name__ == "__main__":
    sys.exit(main())
